"""
Bovino

Usage:
  bovino --template=<template_file> [--mapfile=<mapfile>]  --referto=<referto_file>  [ --filecompilato=<filecompilato> ] 

"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime as _dt
from docopt import docopt 
import logging
import pathlib
import re
logging.basicConfig(
    filename='bovino.log', 
    level=logging.INFO, 
    format='%(asctime)s  {%(pathname)s:%(lineno)d} :: %(levelname)s ->  %(message)s')

class Bovino:

    log = None
    map_data_T1 = {}
    map_data_T2 = {}

    referto_unit_regex = re.compile('(.*?)\s+\((\S*)\)$')

    # we need to copy some metadata in the template
    # their name 
    referto_fields_names = {}
    # and their position inside the xlsx
    referto_fields_pos = {}

    # actually the referto data as rowsz
    referto_data = []
    # all the meta data names available inside the referto xlsx 
    # row #index ---> parameter name
    referto_meta_data = {}

    # the loaded template as xlxs structure
    template_data = None

    # the workbook to save where the above data belongs 
    template_xlsx_to_save = None 

    # sample point name => column index 
    sample_point_pointers = {}



    def __init__(self, ftemplate=None, fmap=None, freferto=None, fcompilato=None) -> None:
        if fcompilato is None:
            p = pathlib.Path(freferto)
            self.fcompilato = p.parent /  pathlib.Path( p.stem + "_compilato_" + p.suffix)
        else:
            self.fcompilato = fcompilato
            
        self.map_file_name = fmap or "MAPPA.xlsx"
        self.log = logging.getLogger()
        self.ref_file_name = freferto or  "SAMPLES/Massivo_20220207-18694.xlsx"
        self.template_file_name = ftemplate or "SAMPLES/59326_Acque_TEMPLATE.xlsx"
        self.log.warning("-" * 80)
        self.log.warning("BOVINO - started")
        self.log.warning("Alessio Palma 2022, released under GNU/GPL see COPYING.txt")
        self.log.warning("-" * 80)
        self.log.info(f"File template : {self.template_file_name}")
        self.log.info(f"File referto  : {self.ref_file_name}")
        self.log.info(f"File mappa : {self.map_file_name}")
        self.log.info(f"File compilato : {self.fcompilato}")

    def load_map(self) -> None:
        self.log.info(f"Sto aprendo il file della mappatura, {self.map_file_name}")
        mapdata = load_workbook(filename=self.map_file_name)
        
        if len(mapdata.sheetnames) != 2:
            raise RuntimeError("Il file della mappatura deve contenere due fogli")
        else:
            if mapdata.sheetnames[0] not in ("Mappa-template-referto", "Intestazioni-referto") \
                and mapdata.sheetnames[1] not in ("Mappa-template-referto", "Intestazioni-referto") \
                and mapdata.sheetnames[0] != mapdata.sheetnames[1]:
                raise RuntimeError("Il file di mappa dovrebbe deve avere 2 fogli, uno chiamato Mappa-template-referto e l'altro Intestazioni-referto")
       
       
        nrow = 1
        ndouble_unit = 0 
        nempty_ref_name = 0
        nref = 0

        for row in mapdata['Mappa-template-referto'].iter_rows():
            template_name = row[0].value 
            template_unit = row[1].value
            ref_name_and_unit = row[2].value
            conversion = row[3].value
            priority = row[4].value

            if ref_name_and_unit is not None:
                self.map_data_T2[ref_name_and_unit] = template_name
                nref += 1

            if template_name is None:
                raise RuntimeError(f"Un nome nel template é non definito, controlla la riga {nrow}")

            if template_name not in self.map_data_T1:
                self.map_data_T1[template_name] =  {  
                    template_unit : { 
                        "unit": template_unit, 
                        "conversion": conversion, 
                        "priority": priority 
                    }  
                }

            else:
                if template_unit in self.map_data_T1[template_name]:
                    raise RuntimeError(f"C'e' un problema con {template_name}, é giá presente una riga con questo valore e questa unitá [{template_unit}] controlla la riga {nrow}")       
                self.map_data_T1[template_name][template_unit] = { 
                        "unit": template_unit, 
                        "conversion" : conversion, 
                        "priority" : priority 
                    }  

                ndouble_unit += 1
            if ref_name_and_unit is None:
                nempty_ref_name += 1

            nrow += 1

        for row in mapdata["Intestazioni-referto"].iter_rows():
            # all the metadata name available inside the referto
            # row index ===> parameter name
            self.referto_meta_data[row[0].row] = row[0].value
            if row[1].value.upper() == 'IGNORA':
                continue
            else:
                self.referto_fields_names[row[0].value]=row[1].value

        self.log.info(f"Ho trovato {nrow} righe")
        self.log.info(f"di queste {nempty_ref_name} non avevano un nome di referto associato")
        self.log.info(f"e {ndouble_unit} sono presenti con lo stesso nome, ma unitá differenti.")
        self.log.info(f"Ho trovato {nref} nomi per le voci dei referti.")
        
        self.log.info("Dal file del referto utilizzeró le seguenti righe di intestazione:")
        for k,v in self.referto_fields_names.items():
            self.log.info(f" {k} ==> {v} ")

    def load_template(self) -> None:
        """loads the template and nothing more"""
        self.log.info(f"Apro il file del template {self.template_file_name}")
        template = load_workbook(filename=self.template_file_name)
        if len(template.sheetnames) != 1:
            raise RuntimeError("Il file del referto puó contenere solo un foglio, questo ne ha piú di uno.")

        wb = template.active
        self.template_data = wb
        self.template_xlsx_to_save = template

        # Checks if the metadata we need are inside the template
        name_and_unit = [ ( row[0].value, row[2].value ) for row in wb.iter_rows()]
        self.log.info(f"Found {len(name_and_unit)} rows")
        name_and_unit = [ a+b for a,b in name_and_unit]

        s = set(name_and_unit)
        if len(name_and_unit) != len(s):
            name_and_unit.sort()
            for i in range(1,len(name_and_unit)):
                if name_and_unit[i-1] == name_and_unit[i]:
                    self.log.error(f"Duplicato: {name_and_unit[i]}")
            raise RuntimeError("Ci sono delle colonne duplicate, cambiare nomi o unitá oppure cancellarle")

        for c in wb.iter_cols():
            self.sample_point_pointers[c[0].value] = c[0].column

    def process_refert_column0(self, col: any) -> None:
        """
        collects the information about the labels and the index where the information to be collect
        are placed
        """
        self.log.info("Processo la colonna zero del referto, quella che contiene le label")
        nnull = 0 
        nrow = 0 
        limit = 1000000

        rows_with_metadata = self.referto_meta_data.keys()
        self.referto_meta_data = {}

        try:
            while ( nnull < 10 ) and ( limit > 0):
                if col[nrow].value is None:
                    nnull += 1
                else:
                    nnull = 0 
                    if col[nrow].value in self.referto_fields_names:
                        self.referto_fields_pos[col[nrow].value] = nrow
                        self.log.info(f"Trovato nel file di referto {col[nrow].value} alla riga {nrow}")
                    # if the current label is not a useless metadata 
                    # save position and names
                    if col[nrow].row not in rows_with_metadata:
                         self.referto_meta_data[nrow] = col[nrow].value
                nrow += 1
                limit -= 1
        except IndexError:
            self.log.info(f"Fine righe. Sono presenti {nrow} righe con valori")
        if len(self.referto_fields_names) != len(self.referto_fields_pos):
            missings = set(self.referto_fields_names.keys()) - set(self.referto_fields_pos.keys()) 
            self.log.error(f"Non ho trovato queste righe nel file di referto: {missings}")
            raise RuntimeError("Non ho trovato alcune righe che dovevano essere presenti, controlla il log per maggiori informazioni")


    def load_referto(self) -> None:
        self.log.info(f"Apro il file del referto {self.ref_file_name}")
        refdata = load_workbook(filename=self.ref_file_name)
        if len(refdata.sheetnames) != 1:
            raise RuntimeError("Il file del referto puó contenere solo un foglio, questo ne ha piú di uno.")
            
        self.log.info("Scansiono il file per colonne")
        ncol = 0
        data = []
        for col in refdata.active.iter_cols():
            if ncol == 0:
                self.process_refert_column0(col)
            else:
                data = []
                # after process_refert_column0 we have the row coords required to process the data 
                for name,row in self.referto_fields_pos.items():
                    value = col[row].value
                    self.log.info(f"Colonna {ncol}, metadato {name} in riga {row} --> {value}")
                    data.append(value)
                for rowindex, name in self.referto_meta_data.items():
                    self.log.info(f"Adding {col[rowindex].value} value for {name}")
                    data.append({  name : col[rowindex].value })
                self.referto_data.append(data)
            ncol += 1
            
        self.log.info("Dati raccolti dal referto:") 
        self.log.info(self.referto_data)
      


    def fill_template(self) -> None:
        sample_points_from_referto = [ x[1] for x in self.referto_data ]
        self.log.info("Verifico che il template abbia tutti i punti di campionamento che sono nel referto")
        
        nsample_points_found = 0
        # Codice punto is supposed to be at Row 0
        for col in self.template_data.iter_cols():
            if col[0].value is None:
                self.log.warning(f"Sto cercando i punti di campionamento ed ho trovato una colonna vuota")
                self.log.warning(f"Colonna vuota: {col[0].column}")
                self.log.warning(f"Termino la ricerca dei punti di campionamento.")
                self.log.warning(f"Eliminare la colonna vuota potrebbe essere una buona idea.")
                break
            if col[0].value not in sample_points_from_referto and len(col[0].value) > 7:
                continue
            else:
                # it is valid sample point code on the template, let's compare with the referto's ones
                self.log.info(f"Ho trovato il punto di campionamento {col[0].value} alla colonna {col[0]}")
                nsample_points_found += 1 
                if col[0].value not in sample_points_from_referto:
                    self.log.error('Questo punto non é presente sul referto')
        if  nsample_points_found != len(sample_points_from_referto):
            raise RuntimeError("Il numero di punti di campionamento sul template non corrisponde a quelli del referto")
        
        chemicals_pointers = {}
        ### collects the row indexes for the chemicals
        for row in self.template_data.iter_rows():
            if row[0].value not in chemicals_pointers:
                # name -> { unit : row } 
                chemicals_pointers[row[0].value] = { row[2].value : row[0].row }
            elif row[2].value in chemicals_pointers[row[0].value]:
                raise RuntimeError(f"Ho trovato {row[0].value} con l'unitá {row[2].value} due volte, significa che ci sono righe duplicate nel template al netto del codice dell'analita, canellale o cambia (temporaneamente) il nome")
            else:
                chemicals_pointers[ row[2].value ] = row[0].row

        ###
        ### Start filling the template for real
        ###
        for d in self.referto_data:
            self.log.info(d)
            self.log.info(f"Sto processando i dati del {d[0]} dal referto, punto di campionamento {d[1]}")
            sample_date = d[0]
            sample_point = d[1]

            for chem_data in d[2:]:
                # ugly but working
                chemical, chem_value = tuple(chem_data.items())[0]
                matches = self.referto_unit_regex.match(chemical)
                # THIS CHEM_NAME is from the referto, we need to convert it into the template name 
                chem_name = matches.group(1)
                tchem_name = self.map_data_T2[chemical]

                chem_unit = matches.group(2)
                chem_unit = "µg/L" if chem_unit == "µg/l" else chem_unit
                chem_unit = "mg/L" if chem_unit == "mg/l" else chem_unit
                chem_unit = "g/L" if chem_unit == "g/l" else chem_unit

                self.log.info(f"--> dati dal referto: {chem_name} con unitá {chem_unit}")
                
                if tchem_name not in chemicals_pointers:
                    raise RuntimeError(f"Non ho trovato {chem_name} all'interno del template, aggiorna il file della mappa")
                else:
                    # the template chemical name from the map is inside the current template
                    # this line collects all the conversion options for this chemical.
                    current_unit_options_for_chem_name = self.map_data_T1.get(tchem_name)

                if chem_unit not in chemicals_pointers[tchem_name]:
                    required_unit = list(chemicals_pointers[tchem_name].keys())
                    self.log.warning(f"{chem_unit} non trovato nel file template, che invece ha {required_unit}")
                    self.log.warning(f"Provo a vedere se sono disponibili conversioni nel file di mappa da da {chem_unit} a {required_unit}")
     
                    if chem_unit in current_unit_options_for_chem_name:
                        self.log.info(f"L'unitá di misura {chem_unit} ==> {required_unit} é presente nelle opzioni di conversione.")                    

                        print(f"Punto di campionamento {sample_point} : {tchem_name} [{chem_unit}] non é presente nel file della mappa. Ci sono queste opzioni:")
                        
                        # quickly check if all the options have no conversioni
                        #
                        conversion_need_options = [ (unit_available, unit_data['conversion'], unit_data['priority']) for unit_available, unit_data in current_unit_options_for_chem_name.items() if unit_data.get('conversion') is not None]
                        
                        if len(conversion_need_options) == 0:
                            self.log.info(f"Non ci sono delle unita che richiedono una conversione. Scelgo in base alla preferenza")                            
                            conversion_need_options = [ (unit_available, unit_data['priority']) for unit_available, unit_data in current_unit_options_for_chem_name.items()]
                            conversion_need_options.sort( key = lambda x : x[1])
                            self.log.info(f"{conversion_need_options}")
                            new_chem_unit = conversion_need_options[0][0]
                            self.log.info(f"Converto {chem_unit} in {new_chem_unit}")
                            chem_unit = new_chem_unit
                        else:
                            self.log.info(f"Purtroppo ci sono delle unita che richiedono una conversione di formato {conversion_need_options}")
                            self.log.info("É tempo che l'utente scelga.")
                        
                            inp = None 
                            while inp is None:
                                c = 0
                                s = {}
                                for unit_available, unit_data in current_unit_options_for_chem_name.items():
                                    if unit_available == chem_unit:
                                        #what we need is not available, skip that option.
                                        continue

                                    print(f" {c} ---> {unit_available} ----------------")
                                    if unit_data.get('priority') is not None :
                                        print(f"    priorità : {unit_data['priority']}")
                                        print(f"    Convesione: {unit_data.get('conversion')}")
                                    s[c] = unit_available
                                    c += 1
                                inp = input(f"Scegli un' unitá sostitutiva (0 .. {c-1}) : ")
                            
                                if not 0 <= int(inp) < c:
                                    inp = None
                                    continue 
                                else:
                                    inp = int(inp)

                                self.log.info(f"L'utente ha scelto:{s[inp]}.")
                                a, b = str(chem_value).split(" ")
                                b = float(b)
                                if unit_data.get('conversion') is not None:
                                    b = b * unit_data.get('conversion')
                                    self.log.info(f"Valore dopo conversione di unitá : {b} ")
                                chem_value = a +" "+str(b)
                                chem_unit = s[inp]


                location_row = chemicals_pointers[tchem_name].get(chem_unit)
                location_column = self.sample_point_pointers[sample_point]
                self.log.info(f"{chem_name}-->{tchem_name} {chem_value} {chem_unit} row={location_row} column={location_column}")
                
                self.template_data.cell(row=2, column=location_column).value = sample_date
                self.template_data.cell(row=location_row, column=location_column).value = chem_value

        self.template_xlsx_to_save.save(filename = self.fcompilato)


    def banner(self,) -> None:
        print("""
        \|/          (__)    
             `\------(oo)           BOVINOMATIC
               ||    (__)
               ||w--||     \|/
        \|/
  ______________________________________________________________________      
  Avviarlo usando WindowsPowerShell app.
  Se necessario attivare l'ambiente con: .\\venv\\Scripts\\Activate.ps1
        """)

    def start(self) -> None:   
        self.banner()
        self.load_map()
        self.load_template()
        self.load_referto()
        self.fill_template()

        
cmdline = "--template SAMPLES\\template_unita_differente.xlsx --referto .\\SAMPLES\\Massivo_20220207-18694.xlsx".split(" ")
opts = docopt(__doc__, argv=cmdline, version=1.0)

ftemplate = opts["--template"] 
fmap = opts["--mapfile"] 
freferto = opts["--referto"]
fcompilato = opts["--filecompilato"]

bovino = Bovino(ftemplate=ftemplate, fmap=fmap, freferto=freferto, fcompilato=fcompilato)
bovino.start()