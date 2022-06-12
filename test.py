from tkinter.messagebox import NO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime as _dt
import docopt 
import logging
logging.basicConfig(filename='log.log', level=logging.INFO)

class Secretary:

    log = None
    map_data_T1 = {}
    map_data_T2 = {}

    # we need to copy some metadata in the template
    # their name 
    referto_fields_names = {}
    # and their position inside the xlsx
    referto_fields_pos = {}

    # actually the referto data as rows
    referto_data = []
    # all the meta data names available inside the referto xlsx 
    # row #index ---> parameter name
    referto_meta_data = {}

    # the loaded template as xlxs structure
    template_data = None

    def __init__(self) -> None:
        self.map_file_name = "MAPPA.xlsx"
        self.log = logging.getLogger()
        self.ref_file_name = "SAMPLES/Massivo_20220207-18694.xlsx"
        self.template_file_name = "SAMPLES/59326_Acque_TEMPLATE.xlsx"


    def load_map(self) -> None:
        self.log.info(f"Sto aprendo il file della mappatura, {self.map_file_name}")
        mapdata = load_workbook(filename=self.map_file_name)
        
        if len(mapdata.sheetnames) != 2:
            raise RuntimeError("Il file della mappatura deve contenere due fogli")
        else:
            if mapdata.sheetnames[0] not in ("Mappa-template-referto", "Intestazioni-referto") \
                and mapdata.sheetnames[1] not in ("Mappa-template-referto", "Intestazioni-referto") \
                and mapdata.sheetnames[0] != mapdata.sheetnames[1]:
                raise RuntimeError("Il file di mappa dovrebbe deve avere 2 fogli, uno chiamato Mappa-template-referto e l'altro Intestazioni-referto")
       
       
        nrow = 1
        ndouble_unit = 0 
        nempty_ref_name = 0
        nref = 0

        for row in mapdata['Mappa-template-referto'].iter_rows():
            template_name = row[0].value 
            template_unit = row[1].value
            ref_name_and_unit = row[2].value
            conversion = row[3].value
            priority = row[4].value

            if ref_name_and_unit is not None:
                self.map_data_T2[ref_name_and_unit] = template_name
                nref += 1

            if template_name is None:
                raise RuntimeError(f"Un nome nel template é non definito, controlla la riga {nrow}")

            if template_name not in self.map_data_T1:
                self.map_data_T1[template_name] =  {  
                    template_unit : { 
                        "unit": template_unit, 
                        "conversion" : conversion, 
                        "priority" : priority 
                    }  
                }

            else:
                if template_unit in self.map_data_T1[template_name]:
                    raise RuntimeError(f"C'e' un problema con {template_name}, é giá presente una riga con questo valore e questa unitá [{template_unit}] controlla la riga {nrow}")       
                self.map_data_T1[template_name][template_unit] = { 
                        "unit": template_unit, 
                        "conversion" : conversion, 
                        "priority" : priority 
                    }  

                ndouble_unit += 1
            if ref_name_and_unit is None:
                nempty_ref_name += 1

            nrow += 1

        for row in mapdata["Intestazioni-referto"].iter_rows():
            # all the metadata name available inside the referto
            # row index ===> parameter name
            self.referto_meta_data[row[0].row] = row[0].value
            if row[1].value.upper() == 'IGNORA':
                continue
            else:
                self.referto_fields_names[row[0].value]=row[1].value

        self.log.info(f"Ho trovato {nrow} righe")
        self.log.info(f"di queste {nempty_ref_name} non avevano un nome di referto associato")
        self.log.info(f"e {ndouble_unit} sono presenti con lo stesso nome, ma unitá differenti.")
        self.log.info(f"Ho trovato {nref} nomi per le voci dei referti.")
        
        self.log.info("Dal file del referto utilizzeró le seguenti righe di intestazione:")
        for k,v in self.referto_fields_names.items():
            self.log.info(f" {k} ==> {v} ")

    def load_template(self) -> None:
        """loads the template and nothing more"""
        self.log.info(f"Apro il file del template {self.template_file_name}")
        template = load_workbook(filename=self.template_file_name)
        if len(template.sheetnames) != 1:
            raise RuntimeError("Il file del referto puó contenere solo un foglio, questo ne ha piú di uno.")

        wb = template.active
        self.template_data = wb

        # Checks if the metadata we need are inside the template
        name_and_unit = [ ( row[0].value, row[2].value ) for row in wb.iter_rows()]
        self.log.info(f"Found {len(name_and_unit)} rows")
        name_and_unit = [ a+b for a,b in name_and_unit]

        s = set(name_and_unit)
        if len(name_and_unit) != len(s):
            name_and_unit.sort()
            for i in range(1,len(name_and_unit)):
                if name_and_unit[i-1] == name_and_unit[i]:
                    self.log.error(f"Duplicato: {name_and_unit[i]}")
            raise RuntimeError("Ci sono delle colonne duplicate, cambiare nomi o unitá oppure cancellarle")


    def process_refert_column0(self, col: any) -> None:
        """
        collects the information about the labels and the index where the information to be collect
        are placed
        """
        self.log.info("Processo la colonna zero del referto, quella che contiene le label")
        print(dir(col))
        nnull = 0 
        nrow = 0 
        limit = 1000000

        rows_with_metadata = self.referto_meta_data.keys()
        self.referto_meta_data = {}

        try:
            while ( nnull < 10 ) and ( limit > 0):
                if col[nrow].value is None:
                    nnull += 1
                else:
                    nnull = 0 
                    if col[nrow].value in self.referto_fields_names:
                        self.referto_fields_pos[col[nrow].value] = nrow
                        self.log.info(f"Trovato nel file di referto {col[nrow].value} alla riga {nrow}")
                    # if the current label is not a useless metadata 
                    # save position and names
                    if col[nrow].row not in rows_with_metadata:
                         self.referto_meta_data[nrow] = col[nrow].value
                nrow += 1
                limit -= 1
        except IndexError:
            self.log.info(f"Fine righe. Sono presenti {nrow} righe con valori")
        if len(self.referto_fields_names) != len(self.referto_fields_pos):
            missings = set(self.referto_fields_names.keys()) - set(self.referto_fields_pos.keys()) 
            self.log.error(f"Non ho trovato queste righe nel file di referto: {missings}")
            raise RuntimeError("Non ho trovato alcune righe che dovevano essere presenti, controlla il log per maggiori informazioni")


    def load_referto(self) -> None:
        self.log.info(f"Apro il file del referto {self.ref_file_name}")
        refdata = load_workbook(filename=self.ref_file_name)
        if len(refdata.sheetnames) != 1:
            raise RuntimeError("Il file del referto puó contenere solo un foglio, questo ne ha piú di uno.")
            
        self.log.info("Scansiono il file per colonne")
        ncol = 0
        data = []
        for col in refdata.active.iter_cols():
            if ncol == 0:
                self.process_refert_column0(col)
            else:
                data = []
                # after process_refert_column0 we have the row coords required to process the data 
                for name,row in self.referto_fields_pos.items():
                    value = col[row].value
                    self.log.info(f"Colonna {ncol}, metadato {name} in riga {row} --> {value}")
                    data.append(value)
                for rowindex, name in self.referto_meta_data.items():
                    print(rowindex, name)

                    self.log.info(f"Adding {col[rowindex].value} value for {name}")
                    data.append({  name : col[rowindex].value })
                self.referto_data.append(data)
            ncol += 1
            
        self.log.info("Dati raccolti dal referto:") 
        self.log.info(self.referto_data)
      


    def fill_template(self) -> None:
        sample_points_from_referto = [ x[1] for x in self.referto_data ]
        self.log.info("Verifico che il template abbia tutti i punti di campionamento che sono nel referto")
        
        nsample_points_found = 0
        # Codice punto is supposed to be at Row 0
        for col in self.template_data.iter_cols():
            if col[0].value not in sample_points_from_referto and len(col[0].value) > 7:
                continue
            else:
                # it is valid sample point code on the template, let's compare with the referto's ones
                self.log.info(f"Ho trovato il punto di campionamento {col[0].value} alla colonna {col[0]}")
                nsample_points_found += 1 
                if col[0].value not in sample_points_from_referto:
                    self.log.error('Questo punto non é presente sul referto')
        if  nsample_points_found != len(sample_points_from_referto):
            raise RuntimeError("Il numero di punti di campionamento sul template non corrisponde a quelli del referto")
        



    def start(self) -> None:
        self.load_map()
        self.load_template()
        self.load_referto()
        self.fill_template()

        



a = Secretary()
a.start()

# wb = load_workbook(filename = './SAMPLES/59326_Acque_TEMPLATE.xlsx')
# sheet_ranges = wb['2_59326']
# print("Lettura dal primo file (a18) ")
# print(sheet_ranges['A18'].value)


# wb2 = load_workbook(filename = 'SAMPLES/Massivo_20220207-18694.xlsx')
# sheet_ranges2 = wb2['Export']
# print("Lettura dal 2nd file (a18) ")
# print(sheet_ranges2['A18'].value)
# print(sheet_ranges2['B1'].value)
# print("Conversione data")
# print( _dt.strptime(sheet_ranges2['B6'].value, "%d/%m/%Y" ) )
